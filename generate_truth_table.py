#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fsQCA 真值表生成与分析脚本
模型: fCDL = f(fGP, fTS, fIF, fTI, fDS, fEN)
算法: Quine-McCluskey

使用方法:
    python generate_truth_table.py

输入文件: SEM_fsQCA_数据_已调整.xlsx
输出文件: fsQCA_truth_table.csv, fsQCA_result_final.csv
"""

import math
import csv
import itertools
from collections import defaultdict

import openpyxl


# ============================================================
# 1. 直接校准法 (Direct Calibration)
# ============================================================

def calibrate_direct(value, fully_out, crossover, fully_in):
    """
    fsQCA 3.0 标准直接校准法 (Ragin 2008)。
    使用分段 logistic 函数，以 ln(19) 为缩放因子，
    确保 fully_out → 0.05, crossover → 0.5, fully_in → 0.95。
    """
    if value <= fully_out:
        return 0.05
    if value >= fully_in:
        return 0.95
    if abs(value - crossover) < 1e-9:
        return 0.5

    # ln(19) ≈ 2.944439, 保证边界映射: 0.05 和 0.95
    LN19 = math.log(19.0)

    if value > crossover:
        k = LN19 / (fully_in - crossover)
    else:
        k = LN19 / (crossover - fully_out)

    score = 1.0 / (1.0 + math.exp(-k * (value - crossover)))
    return score


# ============================================================
# 2. 读取数据
# ============================================================

def read_data(filepath):
    """从 Excel 文件读取原始数据和校准锚点。"""
    wb = openpyxl.load_workbook(filepath, read_only=True)

    # 读取校准锚点
    ws_anchors = wb['fsQCA校准锚点']
    anchors = {}
    for i, row in enumerate(ws_anchors.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳过标题行
        var_name = row[0]
        fully_out = float(row[1])
        crossover_pt = float(row[2])
        fully_in_pt = float(row[3])
        anchors[var_name] = (fully_out, crossover_pt, fully_in_pt)

    # 读取原始数据
    ws_data = wb['fsQCA分析数据(6条件)']
    headers = None
    raw_data = []
    for i, row in enumerate(ws_data.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) for h in row]
            continue
        raw_data.append([float(v) if v is not None else 0.0 for v in row])

    wb.close()
    return headers, raw_data, anchors


# ============================================================
# 3. 校准数据
# ============================================================

def calibrate_data(headers, raw_data, anchors):
    """将原始数据校准为模糊集隶属度。"""
    conditions = ['GP', 'TS', 'IF', 'TI', 'DS', 'EN']
    outcome = 'CDL'
    all_vars = conditions + [outcome]

    # 找到各变量在 headers 中的索引
    var_idx = {}
    for var in all_vars:
        for j, h in enumerate(headers):
            if h == var:
                var_idx[var] = j
                break

    calibrated = []
    for row in raw_data:
        cal_row = {}
        for var in all_vars:
            raw_val = row[var_idx[var]]
            fo, co, fi = anchors[var]
            cal_row['f' + var] = calibrate_direct(raw_val, fo, co, fi)
        calibrated.append(cal_row)

    return calibrated


# ============================================================
# 4. 生成真值表
# ============================================================

def generate_truth_table(calibrated_data, conditions, outcome, freq_cutoff=3):
    """
    生成真值表。
    每个案例根据模糊隶属度 > 0.5 分配到对应的真值表行。
    """
    f_conditions = ['f' + c for c in conditions]
    f_outcome = 'f' + outcome

    # 2^n 种组合
    n = len(conditions)
    truth_table = {}

    for combo in itertools.product([0, 1], repeat=n):
        key = combo
        truth_table[key] = {
            'conditions': dict(zip(conditions, combo)),
            'cases': [],
            'number': 0,
        }

    # 分配案例到真值表行
    for case in calibrated_data:
        row_key = tuple(1 if case[fc] > 0.5 else 0 for fc in f_conditions)
        # 计算该案例在该行的隶属度 (模糊与运算 = 最小值)
        memberships = [case[fc] if combo_val == 1 else (1 - case[fc])
                       for fc, combo_val in zip(f_conditions, row_key)]
        row_membership = min(memberships)
        truth_table[row_key]['cases'].append({
            'membership': row_membership,
            'outcome': case[f_outcome],
        })
        truth_table[row_key]['number'] += 1

    # 计算一致性和覆盖度
    total_outcome_sum = sum(case[f_outcome] for case in calibrated_data)

    results = []
    for key, info in sorted(truth_table.items(), key=lambda x: -x[1]['number']):
        row = dict(zip(conditions, key))
        row['number'] = info['number']

        if info['number'] > 0:
            # 一致性 = Σmin(Xi, Yi) / ΣXi
            sum_min_xy = sum(min(c['membership'], c['outcome']) for c in info['cases'])
            sum_x = sum(c['membership'] for c in info['cases'])
            consistency = sum_min_xy / sum_x if sum_x > 0 else 0

            # PRI (Proportional Reduction in Inconsistency)
            sum_min_xy_neg = sum(min(c['membership'], 1 - c['outcome']) for c in info['cases'])
            pri = (sum_min_xy - sum_min_xy_neg) / (sum_x - sum_min_xy_neg) if (sum_x - sum_min_xy_neg) > 0 else 0

            row['raw_consist'] = round(consistency, 6)
            row['PRI_consist'] = round(pri, 6)
        else:
            row['raw_consist'] = ''
            row['PRI_consist'] = ''

        results.append(row)

    return results


# ============================================================
# 5. Quine-McCluskey 布尔最小化
# ============================================================

def get_prime_implicants(minterms, n_vars):
    """Quine-McCluskey 算法求质蕴涵项。"""
    # 初始化：将每个最小项转为二进制字符串
    groups = defaultdict(set)
    for m in minterms:
        bits = format(m, f'0{n_vars}b')
        ones = bits.count('1')
        groups[ones].add((bits, frozenset([m])))

    prime_implicants = set()
    used = set()

    while groups:
        new_groups = defaultdict(set)
        used_in_round = set()

        sorted_keys = sorted(groups.keys())
        for i in range(len(sorted_keys) - 1):
            k1, k2 = sorted_keys[i], sorted_keys[i + 1]
            if k2 - k1 != 1:
                continue
            for bits1, mterms1 in groups[k1]:
                for bits2, mterms2 in groups[k2]:
                    # 检查是否只有一位不同
                    diff_pos = [j for j in range(n_vars) if bits1[j] != bits2[j]]
                    if len(diff_pos) == 1:
                        new_bits = bits1[:diff_pos[0]] + '-' + bits1[diff_pos[0] + 1:]
                        combined = mterms1 | mterms2
                        new_ones = new_bits.replace('-', '').count('1')
                        new_groups[new_ones].add((new_bits, combined))
                        used_in_round.add((bits1, mterms1))
                        used_in_round.add((bits2, mterms2))

        # 未被合并的项是质蕴涵项
        for k in groups:
            for item in groups[k]:
                if item not in used_in_round:
                    prime_implicants.add(item)

        groups = new_groups

    return prime_implicants


def find_essential_prime_implicants(prime_implicants, minterms):
    """找出必要质蕴涵项和最小覆盖。"""
    # 建立覆盖表
    pi_list = list(prime_implicants)

    # 找出每个最小项被哪些质蕴涵项覆盖
    coverage = defaultdict(list)
    for i, (bits, mterms) in enumerate(pi_list):
        for m in mterms:
            if m in minterms:
                coverage[m].append(i)

    essential = set()
    remaining_minterms = set(minterms)

    # 找必要质蕴涵项：只被一个质蕴涵项覆盖的最小项
    for m in list(remaining_minterms):
        if m in coverage and len(coverage[m]) == 1:
            pi_idx = coverage[m][0]
            essential.add(pi_idx)
            # 移除该质蕴涵项覆盖的所有最小项
            _, covered = pi_list[pi_idx]
            remaining_minterms -= covered

    # 如果还有未覆盖的最小项，使用贪心法
    while remaining_minterms:
        best_pi = None
        best_cover = 0
        for i, (bits, mterms) in enumerate(pi_list):
            if i in essential:
                continue
            cover_count = len(mterms & remaining_minterms)
            if cover_count > best_cover:
                best_cover = cover_count
                best_pi = i
        if best_pi is None:
            break
        essential.add(best_pi)
        _, covered = pi_list[best_pi]
        remaining_minterms -= covered

    return [pi_list[i] for i in essential]


def bits_to_expression(bits, var_names):
    """将二进制/dash 模式转为表达式。"""
    terms = []
    for i, b in enumerate(bits):
        if b == '1':
            terms.append(var_names[i])
        elif b == '0':
            terms.append('~' + var_names[i])
        # '-' 表示不关心
    return '*'.join(terms) if terms else '1'


def bits_to_expression_present_only(bits, var_names):
    """将二进制/dash 模式转为表达式（仅保留 present 条件）。"""
    terms = []
    for i, b in enumerate(bits):
        if b == '1':
            terms.append(var_names[i])
        # '0' 和 '-' 都忽略
    return '*'.join(terms) if terms else '1'


# ============================================================
# 6. 计算覆盖度和一致性
# ============================================================

def compute_solution_metrics(calibrated_data, solution_terms, f_conditions, f_outcome):
    """计算每个解的原始覆盖度、唯一覆盖度和一致性。"""
    n = len(calibrated_data)

    # 计算每个方案的模糊隶属度
    term_memberships = []
    for bits, _ in solution_terms:
        memberships = []
        for case in calibrated_data:
            m = 1.0
            for i, b in enumerate(bits):
                if b == '1':
                    m = min(m, case[f_conditions[i]])
                elif b == '0':
                    m = min(m, 1 - case[f_conditions[i]])
                # '-' 不约束
            memberships.append(m)
        term_memberships.append(memberships)

    # 计算整体解的隶属度
    solution_memberships = []
    for j in range(n):
        solution_memberships.append(max(tm[j] for tm in term_memberships))

    results = []
    outcome_values = [case[f_outcome] for case in calibrated_data]

    for t_idx, (bits, _) in enumerate(solution_terms):
        tm = term_memberships[t_idx]

        # 原始覆盖度 = Σmin(Xi, Yi) / ΣYi
        raw_coverage = sum(min(tm[j], outcome_values[j]) for j in range(n)) / sum(outcome_values)

        # 唯一覆盖度
        # 计算不含当前项的解的隶属度
        other_max = []
        for j in range(n):
            others = [term_memberships[k][j] for k in range(len(term_memberships)) if k != t_idx]
            other_max.append(max(others) if others else 0)

        unique_coverage = sum(
            max(0, min(tm[j], outcome_values[j]) - min(other_max[j], outcome_values[j]))
            for j in range(n)
        ) / sum(outcome_values)

        # 一致性 = Σmin(Xi, Yi) / ΣXi
        sum_x = sum(tm)
        consistency = sum(min(tm[j], outcome_values[j]) for j in range(n)) / sum_x if sum_x > 0 else 0

        results.append({
            'raw_coverage': raw_coverage,
            'unique_coverage': unique_coverage,
            'consistency': consistency,
        })

    # 整体解的覆盖度和一致性
    sol_coverage = sum(min(solution_memberships[j], outcome_values[j]) for j in range(n)) / sum(outcome_values)
    sum_sol = sum(solution_memberships)
    sol_consistency = sum(min(solution_memberships[j], outcome_values[j]) for j in range(n)) / sum_sol if sum_sol > 0 else 0

    return results, sol_coverage, sol_consistency


# ============================================================
# 7. 中间解 (Intermediate Solution)
# ============================================================

def generate_intermediate_solution(prime_implicants, minterms, dc_set, var_names, assumptions,
                                   calibrated_data, f_conditions, f_outcome):
    """
    生成中间解。
    使用简易期望（directional expectations）过滤质蕴涵项。
    assumptions: dict, key=var_name, value='present' or 'absent'
    """
    # 过滤质蕴涵项：移除与期望方向矛盾的项
    filtered_pis = []
    for bits, mterms in prime_implicants:
        valid = True
        for i, b in enumerate(bits):
            if b == '-':
                continue
            var = var_names[i]
            if var in assumptions:
                expected = assumptions[var]
                if expected == 'present' and b == '0':
                    valid = False
                    break
                elif expected == 'absent' and b == '1':
                    valid = False
                    break
        if valid:
            # 只保留覆盖了至少一个实际最小项的质蕴涵项
            if mterms & minterms:
                filtered_pis.append((bits, mterms))

    if not filtered_pis:
        return [], 0, 0

    # 从过滤后的质蕴涵项中找最小覆盖
    essential = find_essential_prime_implicants(set(filtered_pis), minterms)

    # 计算覆盖度和一致性
    metrics, sol_cov, sol_con = compute_solution_metrics(
        calibrated_data, essential, f_conditions, f_outcome
    )

    return essential, metrics, sol_cov, sol_con


# ============================================================
# 8. 主程序
# ============================================================

def main():
    filepath = 'SEM_fsQCA_数据_已调整.xlsx'
    conditions = ['GP', 'TS', 'IF', 'TI', 'DS', 'EN']
    outcome = 'CDL'
    f_conditions = ['f' + c for c in conditions]
    f_outcome = 'f' + outcome
    f_var_names = ['f' + c for c in conditions]
    n_conds = len(conditions)

    print("=" * 60)
    print("fsQCA 真值表分析")
    print("=" * 60)
    print(f"模型: fCDL = f({', '.join(f_var_names)})")
    print(f"算法: Quine-McCluskey")
    print()

    # 1. 读取数据
    print("正在读取数据...")
    headers, raw_data, anchors = read_data(filepath)
    print(f"  案例数: {len(raw_data)}")

    # 2. 校准数据
    print("正在校准数据...")
    calibrated = calibrate_data(headers, raw_data, anchors)

    # 保存校准后的数据
    with open('fsQCA_result_final.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        cal_headers = ['序号'] + f_var_names + [f_outcome]
        writer.writerow(cal_headers)
        for i, case in enumerate(calibrated):
            writer.writerow([i + 1] + [round(case[fc], 6) for fc in f_var_names] +
                            [round(case[f_outcome], 6)])
    print(f"  校准数据已保存至 fsQCA_result_final.csv")

    # 3. 生成真值表
    print("\n正在生成真值表...")
    truth_table = generate_truth_table(calibrated, conditions, outcome)

    # 保存真值表
    tt_headers = conditions + ['number', 'raw_consist', 'PRI_consist']
    with open('fsQCA_truth_table.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=tt_headers)
        writer.writeheader()
        writer.writerows(truth_table)
    print(f"  真值表已保存至 fsQCA_truth_table.csv")

    # 4. 打印真值表
    print(f"\n{'='*60}")
    print("真 值 表 (TRUTH TABLE)")
    print(f"{'='*60}")
    print(f"{'GP':>4} {'TS':>4} {'IF':>4} {'TI':>4} {'DS':>4} {'EN':>4} {'案例数':>6} {'一致性':>10} {'PRI':>10}")
    print("-" * 60)

    freq_cutoff = 3
    for row in truth_table:
        if row['number'] > 0:
            consist_str = f"{row['raw_consist']:.6f}" if isinstance(row['raw_consist'], float) else ''
            pri_str = f"{row['PRI_consist']:.6f}" if isinstance(row['PRI_consist'], float) else ''
            marker = " *" if row['number'] >= freq_cutoff and isinstance(row['raw_consist'], float) and row['raw_consist'] >= 0.85 else ""
            print(f"{row['GP']:>4} {row['TS']:>4} {row['IF']:>4} {row['TI']:>4} {row['DS']:>4} {row['EN']:>4} "
                  f"{row['number']:>6} {consist_str:>10} {pri_str:>10}{marker}")

    # 5. 确定频率和一致性截断值
    # 按一致性排序，找到自然断裂点
    valid_rows = [r for r in truth_table if r['number'] >= freq_cutoff and isinstance(r['raw_consist'], float)]
    valid_rows.sort(key=lambda x: -x['raw_consist'])

    print(f"\n频率截断值 (frequency cutoff): {freq_cutoff}")

    # 找到一致性截断点
    consistencies = sorted(set(r['raw_consist'] for r in valid_rows), reverse=True)
    print(f"有效行的一致性值 (降序): {[f'{c:.6f}' for c in consistencies[:15]]}")

    # 使用 0.85 作为一致性截断值（可调整）
    consistency_cutoff = 0.85

    # 检查问题描述中的截断值
    target_cutoff = 0.871148
    # 找到最接近的自然断裂点
    for i in range(len(consistencies) - 1):
        if consistencies[i] >= target_cutoff > consistencies[i + 1]:
            consistency_cutoff = consistencies[i + 1] if abs(consistencies[i + 1] - target_cutoff) < abs(consistencies[i] - target_cutoff) else consistencies[i]
            break

    # 直接使用目标截断值
    consistency_cutoff = target_cutoff
    print(f"一致性截断值 (consistency cutoff): {consistency_cutoff:.6f}")

    # 6. 确定最小项
    positive_rows = [r for r in valid_rows if r['raw_consist'] >= consistency_cutoff]
    minterms = set()
    for row in positive_rows:
        minterm = 0
        for i, c in enumerate(conditions):
            if row[c] == 1:
                minterm |= (1 << (n_conds - 1 - i))
        minterms.add(minterm)

    # 无关项（don't care）: 频率不足的行
    dc_rows = [r for r in truth_table if r['number'] < freq_cutoff or r['number'] == 0]
    dc_set = set()
    for row in dc_rows:
        minterm = 0
        for i, c in enumerate(conditions):
            if row[c] == 1:
                minterm |= (1 << (n_conds - 1 - i))
        dc_set.add(minterm)

    print(f"\n结果为 1 的行数: {len(positive_rows)}")
    print(f"最小项: {sorted(minterms)}")
    print(f"无关项数: {len(dc_set)}")

    # 7. 简约解 (Parsimonious Solution)
    print(f"\n{'='*60}")
    print("简约解 (PARSIMONIOUS SOLUTION)")
    print(f"{'='*60}")

    all_terms = minterms | dc_set
    prime_implicants = get_prime_implicants(all_terms, n_conds)
    essential_pis = find_essential_prime_implicants(prime_implicants, minterms)

    print(f"\n--- TRUTH TABLE SOLUTION ---")
    print(f"frequency cutoff: {freq_cutoff}")
    print(f"consistency cutoff: {consistency_cutoff}")
    print(f"Assumptions:")

    # 计算度量
    if essential_pis:
        metrics, sol_cov, sol_con = compute_solution_metrics(
            calibrated, essential_pis, f_conditions, f_outcome
        )

        print(f"{'':>20}{'raw':>10}{'unique':>12}")
        print(f"{'':>18}{'coverage':>10}{'coverage':>12}{'consistency':>14}")
        print(f"{'':>18}{'----------':>10}{'----------':>12}{'----------':>14}")

        for i, (bits, _) in enumerate(essential_pis):
            expr = bits_to_expression_present_only(bits, f_var_names)
            m = metrics[i]
            print(f"{expr:>18}{m['raw_coverage']:>10.6f}{m['unique_coverage']:>12.6f}{m['consistency']:>14.6f}")

        print(f"\nsolution coverage: {sol_cov:.6f}")
        print(f"solution consistency: {sol_con:.6f}")
    else:
        print("  无简约解")

    # 8. 中间解 (Intermediate Solution)
    print(f"\n{'='*60}")
    print("中间解 (INTERMEDIATE SOLUTION)")
    print(f"{'='*60}")

    # 所有条件设为 present
    assumptions = {f'f{c}': 'present' for c in conditions}

    # 重新计算中间解的质蕴涵项（只使用最小项，不使用无关项进行扩展）
    # 中间解：使用所有质蕴涵项但过滤方向期望
    intermediate_pis, int_metrics, int_sol_cov, int_sol_con = generate_intermediate_solution(
        prime_implicants, minterms, dc_set, f_var_names, assumptions,
        calibrated, f_conditions, f_outcome
    )

    print(f"\n--- INTERMEDIATE SOLUTION ---")
    print(f"frequency cutoff: {freq_cutoff}")
    print(f"consistency cutoff: {consistency_cutoff}")
    print(f"Assumptions:")
    for c in conditions:
        print(f"f{c} (present)")

    if intermediate_pis:
        print(f"{'':>20}{'raw':>10}{'unique':>12}")
        print(f"{'':>18}{'coverage':>10}{'coverage':>12}{'consistency':>14}")
        print(f"{'':>18}{'----------':>10}{'----------':>12}{'----------':>14}")

        for i, (bits, _) in enumerate(intermediate_pis):
            expr = bits_to_expression_present_only(bits, f_var_names)
            m = int_metrics[i]
            print(f"{expr:>18}{m['raw_coverage']:>10.6f}{m['unique_coverage']:>12.6f}{m['consistency']:>14.6f}")

        print(f"\nsolution coverage: {int_sol_cov:.6f}")
        print(f"solution consistency: {int_sol_con:.6f}")
    else:
        print("  无中间解")

    # 9. 打印所有质蕴涵项（调试用）
    print(f"\n{'='*60}")
    print("所有质蕴涵项 (ALL PRIME IMPLICANTS)")
    print(f"{'='*60}")
    for bits, mterms in sorted(prime_implicants, key=lambda x: x[0].count('-'), reverse=True):
        expr_full = bits_to_expression(bits, f_var_names)
        expr_present = bits_to_expression_present_only(bits, f_var_names)
        print(f"  {bits}  =>  {expr_present:30s}  覆盖最小项: {sorted(mterms)}")


if __name__ == '__main__':
    main()
